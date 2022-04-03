import csv
from django.shortcuts import redirect, render
from django.contrib.admin.views.decorators import staff_member_required
import openpyxl
from cards.models import Country
from cards.models import CouponCard
from cards.models import Batch
from django.contrib import messages
from sellers.models import Users


def handle_parameters_upload(request, file):

    wb = openpyxl.load_workbook(file, read_only=True)
    first_sheet = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(first_sheet)

    data = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        parameter = Country()
        parameter.code = row[0].value
        parameter.name = row[1].value

        data.append(parameter)

    Country.objects.bulk_create(data)

    return True


def handle_parameters_upload_cards(request, file, usser):

    wb = openpyxl.load_workbook(file, read_only=True)
    first_sheet = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(first_sheet)

    data = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):

        country = Country.objects.filter(code=str(row[0].value)[0:6])
        country_name = ""
        for c in country:
            country_name += c.name
        


        user = Users.objects.filter(username=usser).first()
        batchh, created = Batch.objects.update_or_create(
            name=row[3].value, status="Working"
        )
        batchh.save()
        batch = Batch.objects.filter(name=row[3].value).first()
        parameter = CouponCard()
        parameter.serial = row[0].value
        parameter.expiry_date = row[1].value
        parameter.cvv = row[2].value
        parameter.batch = batch
        parameter.seller = user
        parameter.country = str(country_name) or 'Unknown'

        if parameter.country == "UNITED STATES":
            parameter.price = 5
        elif parameter.country != "UNITED STATES" and parameter.country != 'Unknown':
            parameter.price = 10
        else:
            parameter.price = 7

        data.append(parameter)
    CouponCard.objects.bulk_create(data)

    return True


def model_form_upload(request):
    if request.method == "POST" and request.FILES["upload"]:
        upload = request.FILES["upload"]
        if upload:
            x = handle_parameters_upload(request, upload)
            messages.add_message(
                request, messages.SUCCESS, "File Was Uploaded Successfully"
            )
            return redirect("#")
        return render(request, "includes/imported.html", {"": ""})
    return render(request, "includes/imported.html")


def model_form_upload_cards(request):
    if request.method == "POST" and request.FILES["upload-cards"]:
        print(request.user, 'usssseeeeeeeee')
        upload = request.FILES["upload-cards"]
        usser = request.user
        if upload:
            x = handle_parameters_upload_cards(request, upload, usser)
            messages.add_message(
                request, messages.SUCCESS, "File Was Uploaded Successfully"
            )
            return redirect("home")
        return render(request, "includes/imported.html", {"": ""})
    return render(request, "includes/card_import.html")
